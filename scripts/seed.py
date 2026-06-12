#!/usr/bin/env python3
"""Seed Supabase from Excel."""
import os, re
from openpyxl import load_workbook
from supabase import create_client
from dotenv import load_dotenv

load_dotenv(".env.local")

url = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
if not url or not key:
    print("ERROR: Set NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in .env.local")
    exit(1)

supabase = create_client(url, key)

EXCEL = "/Users/dnzyc/texas-apartments-scraper/texas_apartments_with_zip.xlsx"
COL_MAP = {
    0: "name", 1: "address", 2: "city", 3: "phone", 4: "website",
    5: "lat", 6: "lng", 7: "rating", 8: "review_count", 9: "maps_link",
    10: "photo_url", 11: "hours", 12: "category", 13: "region",
    14: "scrape_date", 15: "zip_code", 16: "county"
}

def slugify(name, zip_code):
    s = re.sub(r"[^a-z0-9]+", "-", str(name or "").lower()).strip("-")
    return f"{s}-{zip_code}"

wb = load_workbook(EXCEL)
ws = wb.active
total = ws.max_row - 1

batch = []
slug_counts = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    r = {COL_MAP[i]: row[i] for i in COL_MAP if i < len(row)}

    base_slug = slugify(r["name"], str(r.get("zip_code", "")))
    n = slug_counts.get(base_slug, 0) + 1
    slug_counts[base_slug] = n
    slug = f"{base_slug}-{n}" if n > 1 else base_slug

    lat = float(r["lat"]) if r.get("lat") else None
    lng = float(r["lng"]) if r.get("lng") else None
    location = f"POINT({lng} {lat})" if lat is not None and lng is not None else None

    rating = float(r["rating"]) if r.get("rating") else None
    rc = int(r["review_count"]) if r.get("review_count") else None

    place = {
        "slug": slug,
        "name": r["name"],
        "address": r.get("address"),
        "city": r.get("city"),
        "zip_code": str(r.get("zip_code", "")),
        "county": r.get("county"),
        "region": r.get("region"),
        "location": location,
        "rating": rating,
        "review_count": rc,
        "phone": r.get("phone"),
        "website": r.get("website"),
        "maps_link": r.get("maps_link"),
        "photo_url": r.get("photo_url"),
        "hours": r.get("hours"),
        "category": r.get("category"),
    }

    batch.append(place)

    if len(batch) >= 500:
        try:
            supabase.table("places").insert(batch).execute()
            print(f"  Inserted {len(batch)} rows")
        except Exception as e:
            print(f"  ERROR: {e}")
        batch = []

if batch:
    try:
        supabase.table("places").insert(batch).execute()
        print(f"  Inserted final {len(batch)} rows")
    except Exception as e:
        print(f"  ERROR: {e}")

print(f"\nDone! {total} rows processed.")
